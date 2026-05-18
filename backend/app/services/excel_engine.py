"""
Excel export engine using openpyxl.
Produces a multi-sheet workbook:
  Sheet 1 "Document"   — full document text content
  Sheet 2 "Chart Data" — data tables for every chart spec
  Sheet 3+ "Chart N"   — one sheet per chart with an embedded Excel chart
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.schemas import ChartSpec, DocumentSection


# ── style constants ────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
_ACCENT_FILL  = PatternFill("solid", fgColor="334155")
_BAND_FILL    = PatternFill("solid", fgColor="F8FAFC")
_HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT    = Font(name="Calibri", size=10)
_TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1E293B")
_SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1E293B")
_THIN_BORDER  = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _cell(ws, row: int, col: int, value: Any = None, *,
          font: Font | None = None,
          fill: PatternFill | None = None,
          align: str = "left",
          wrap: bool = False,
          border: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = _THIN_BORDER


# ── Sheet 1: Document content ─────────────────────────────────────────────────

def _build_document_sheet(wb: Workbook, title: str,
                           sections: list[DocumentSection],
                           outline: list[str]) -> None:
    ws = wb.active
    ws.title = "Document"
    ws.sheet_view.showGridLines = False

    # title row
    ws.row_dimensions[1].height = 32
    _cell(ws, 1, 1, title, font=_TITLE_FONT, fill=_HEADER_FILL,
          align="center", border=True)
    ws.merge_cells("A1:C1")

    headers = [("Section #", 12), ("Heading", 36), ("Content", 80)]
    for col, (h, w) in enumerate(headers, start=1):
        _cell(ws, 2, col, h, font=_HEADER_FONT, fill=_HEADER_FILL,
              align="center", border=True)
        _set_col_width(ws, col, w)
    ws.row_dimensions[2].height = 22

    row = 3
    for i, sec in enumerate(sections, 1):
        fill = _BAND_FILL if i % 2 == 0 else None
        _cell(ws, row, 1, i, font=_BODY_FONT, fill=fill,
              align="center", border=True)
        _cell(ws, row, 2, sec.heading, font=_SECTION_FONT, fill=fill,
              border=True, wrap=True)
        _cell(ws, row, 3, sec.content or "", font=_BODY_FONT, fill=fill,
              border=True, wrap=True)
        # auto-height: roughly 15 pts per ~100 chars
        content_len = len(sec.content or "")
        ws.row_dimensions[row].height = max(40, min(120, 15 + content_len // 6))
        row += 1

    # freeze header rows
    ws.freeze_panes = "A3"


# ── Sheet 2: Chart data table ──────────────────────────────────────────────────

def _build_chart_data_sheet(wb: Workbook, specs: list[ChartSpec]) -> None:
    ws = wb.create_sheet("Chart Data")
    ws.sheet_view.showGridLines = False

    col = 1
    for idx, spec in enumerate(specs, 1):
        # chart title
        title_cell = f"{get_column_letter(col)}1"
        _cell(ws, 1, col, f"Chart {idx}: {spec.title}",
              font=_HEADER_FONT, fill=_HEADER_FILL, align="center", border=True)
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + 1
        )

        # headers
        _cell(ws, 2, col,   spec.x_label or "Label",
              font=_HEADER_FONT, fill=_ACCENT_FILL, align="center", border=True)
        _cell(ws, 2, col+1, spec.y_label or "Value",
              font=_HEADER_FONT, fill=_ACCENT_FILL, align="center", border=True)
        _set_col_width(ws, col,   22)
        _set_col_width(ws, col+1, 16)

        # data rows
        for r, (lbl, val) in enumerate(zip(spec.labels, spec.values), start=3):
            fill = _BAND_FILL if r % 2 == 0 else None
            _cell(ws, r, col,   lbl, font=_BODY_FONT, fill=fill, border=True)
            _cell(ws, r, col+1, val, font=_BODY_FONT, fill=fill,
                  align="right", border=True)

        col += 3  # gap between tables

    ws.freeze_panes = "A3"


# ── Sheet 3+: Per-chart sheets with embedded charts ───────────────────────────

def _build_chart_sheet(wb: Workbook, spec: ChartSpec, idx: int) -> None:
    sheet_name = f"Chart {idx}"[:31]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # data table
    _cell(ws, 1, 1, spec.x_label or "Category",
          font=_HEADER_FONT, fill=_HEADER_FILL, border=True)
    _cell(ws, 1, 2, spec.y_label or "Value",
          font=_HEADER_FONT, fill=_HEADER_FILL, border=True)
    _set_col_width(ws, 1, 22)
    _set_col_width(ws, 2, 16)

    for r, (lbl, val) in enumerate(zip(spec.labels, spec.values), start=2):
        fill = _BAND_FILL if r % 2 == 0 else None
        _cell(ws, r, 1, lbl, font=_BODY_FONT, fill=fill, border=True)
        _cell(ws, r, 2, val, font=_BODY_FONT, fill=fill,
              align="right", border=True)

    n = len(spec.values)
    data_ref  = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    label_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)

    if spec.kind == "bar":
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.overlap = -10
    elif spec.kind == "line":
        chart = LineChart()
        chart.smooth = True
    else:
        chart = PieChart()

    chart.title = spec.title or f"Chart {idx}"
    chart.style = 10
    chart.width  = 18
    chart.height = 12

    chart.add_data(data_ref, titles_from_data=True)

    if spec.kind != "pie":
        chart.set_categories(label_ref)
        if spec.x_label:
            chart.x_axis.title = spec.x_label
        if spec.y_label:
            chart.y_axis.title = spec.y_label
    else:
        chart.set_categories(label_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True

    ws.add_chart(chart, "D2")


# ── public entry point ────────────────────────────────────────────────────────

def build_excel(
    *,
    out_path: Path,
    title: str,
    outline: list[str],
    sections: list[DocumentSection],
    chart_specs: Optional[list[ChartSpec]] = None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    specs = chart_specs or []

    _build_document_sheet(wb, title, sections, outline)

    if specs:
        _build_chart_data_sheet(wb, specs)
        for i, spec in enumerate(specs, 1):
            _build_chart_sheet(wb, spec, i)

    wb.save(str(out_path))
    return out_path
